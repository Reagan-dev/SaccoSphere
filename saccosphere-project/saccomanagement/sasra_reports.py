"""
The exact cell layout, sheet names, and column order in the official SASRA
return template can change between regulatory cycles. This view produces the
correct UNDERLYING figures using SASRA's published classification and
provisioning rules; before submission, the SACCO's finance officer must
confirm the column layout still matches the current official SASRA workbook
and remap headers here if SASRA has revised the template.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from accounts.permissions import IsSaccoAdmin
from ledger.models import LedgerEntry
from saccomanagement.models import SystemAuditLog, Role
from saccomembership.models import Membership
from services.models import Loan, RepaymentSchedule, Saving, SavingsType


# SASRA PAR classification and provisioning rates
PAR_CATEGORIES = {
    'PERFORMING': {
        'days_max': 0,
        'provision_rate': Decimal('0.01'),  # 1%
        'label': 'Performing',
    },
    'WATCH': {
        'days_min': 1,
        'days_max': 30,
        'provision_rate': Decimal('0.03'),  # 3%
        'label': 'Watch',
    },
    'SUBSTANDARD': {
        'days_min': 31,
        'days_max': 90,
        'provision_rate': Decimal('0.20'),  # 20%
        'label': 'Substandard',
    },
    'DOUBTFUL': {
        'days_min': 91,
        'days_max': 180,
        'provision_rate': Decimal('0.50'),  # 50%
        'label': 'Doubtful',
    },
    'LOSS': {
        'days_min': 181,
        'provision_rate': Decimal('1.00'),  # 100%
        'label': 'Loss',
    },
}


def build_par_return(sacco, as_of_date):
    """
    Build Portfolio at Risk (PAR) return for SASRA reporting.

    Classifies ACTIVE loans into SASRA's five risk categories based on
    days overdue of the earliest unpaid instalment, calculates provision
    requirements, and computes PAR30/PAR90 ratios.

    Args:
        sacco: Sacco instance
        as_of_date: date object for the report cutoff

    Returns:
        dict with classification breakdown, totals, and PAR ratios
    """
    # Get all ACTIVE loans for the SACCO
    active_loans = Loan.objects.filter(
        membership__sacco=sacco,
        status=Loan.Status.ACTIVE,
    ).select_related('membership')

    total_outstanding = active_loans.aggregate(
        total=Sum('outstanding_balance')
    )['total'] or Decimal('0.00')

    # Initialize category buckets
    categories = {
        'PERFORMING': {
            'loan_count': 0,
            'outstanding_balance': Decimal('0.00'),
            'provision_required': Decimal('0.00'),
        },
        'WATCH': {
            'loan_count': 0,
            'outstanding_balance': Decimal('0.00'),
            'provision_required': Decimal('0.00'),
        },
        'SUBSTANDARD': {
            'loan_count': 0,
            'outstanding_balance': Decimal('0.00'),
            'provision_required': Decimal('0.00'),
        },
        'DOUBTFUL': {
            'loan_count': 0,
            'outstanding_balance': Decimal('0.00'),
            'provision_required': Decimal('0.00'),
        },
        'LOSS': {
            'loan_count': 0,
            'outstanding_balance': Decimal('0.00'),
            'provision_required': Decimal('0.00'),
        },
    }

    # Classify each loan
    for loan in active_loans:
        # Find earliest unpaid instalment
        earliest_unpaid = loan.schedule.filter(
            status=RepaymentSchedule.Status.PENDING
        ).order_by('due_date').first()

        if earliest_unpaid:
            days_overdue = earliest_unpaid.days_overdue
        else:
            # No unpaid instalments - loan is performing
            days_overdue = 0

        # Determine category
        category = _classify_loan_by_days_overdue(days_overdue)

        # Add to bucket
        categories[category]['loan_count'] += 1
        categories[category]['outstanding_balance'] += loan.outstanding_balance

        # Calculate provision
        provision_rate = PAR_CATEGORIES[category]['provision_rate']
        provision = (loan.outstanding_balance * provision_rate).quantize(
            Decimal('0.01')
        )
        categories[category]['provision_required'] += provision

    # Calculate PAR ratios
    par30_balance = (
        categories['WATCH']['outstanding_balance'] +
        categories['SUBSTANDARD']['outstanding_balance'] +
        categories['DOUBTFUL']['outstanding_balance'] +
        categories['LOSS']['outstanding_balance']
    )

    par90_balance = (
        categories['SUBSTANDARD']['outstanding_balance'] +
        categories['DOUBTFUL']['outstanding_balance'] +
        categories['LOSS']['outstanding_balance']
    )

    par30_ratio = (
        (par30_balance / total_outstanding * 100).quantize(Decimal('0.01'))
        if total_outstanding > 0 else Decimal('0.00')
    )

    par90_ratio = (
        (par90_balance / total_outstanding * 100).quantize(Decimal('0.01'))
        if total_outstanding > 0 else Decimal('0.00')
    )

    return {
        'as_of_date': as_of_date.isoformat(),
        'total_outstanding_book': str(total_outstanding),
        'categories': {
            key: {
                'label': PAR_CATEGORIES[key]['label'],
                'loan_count': value['loan_count'],
                'outstanding_balance': str(value['outstanding_balance']),
                'provision_required': str(value['provision_required']),
            }
            for key, value in categories.items()
        },
        'par30_ratio': str(par30_ratio),
        'par90_ratio': str(par90_ratio),
    }


def _classify_loan_by_days_overdue(days_overdue):
    """Classify loan into SASRA PAR category based on days overdue."""
    if days_overdue <= PAR_CATEGORIES['PERFORMING']['days_max']:
        return 'PERFORMING'
    elif days_overdue <= PAR_CATEGORIES['WATCH']['days_max']:
        return 'WATCH'
    elif days_overdue <= PAR_CATEGORIES['SUBSTANDARD']['days_max']:
        return 'SUBSTANDARD'
    elif days_overdue <= PAR_CATEGORIES['DOUBTFUL']['days_max']:
        return 'DOUBTFUL'
    else:
        return 'LOSS'


def build_financial_position_return(sacco, as_of_date):
    """
    Build simplified statement of financial position for SASRA reporting.

    Assets = loans outstanding (ACTIVE loans) + cash balance
    Liabilities = member savings by type (BOSA/FOSA/SHARE_CAPITAL)

    Cash balance is calculated from LedgerEntry by summing credits minus
    debits for cash-related categories.

    Args:
        sacco: Sacco instance
        as_of_date: date object for the report cutoff

    Returns:
        dict with assets and liabilities breakdown
    """
    # Assets: Loans outstanding
    loans_outstanding = Loan.objects.filter(
        membership__sacco=sacco,
        status=Loan.Status.ACTIVE,
    ).aggregate(total=Sum('outstanding_balance'))['total'] or Decimal('0.00')

    # Assets: Cash balance from LedgerEntry
    # Cash-in categories: SAVING_DEPOSIT, LOAN_REPAYMENT, FEE, PENALTY
    # Cash-out categories: SAVING_WITHDRAWAL, LOAN_DISBURSEMENT, DIVIDEND
    cash_in_categories = [
        LedgerEntry.Category.SAVING_DEPOSIT,
        LedgerEntry.Category.LOAN_REPAYMENT,
        LedgerEntry.Category.FEE,
        LedgerEntry.Category.PENALTY,
    ]

    cash_out_categories = [
        LedgerEntry.Category.SAVING_WITHDRAWAL,
        LedgerEntry.Category.LOAN_DISBURSEMENT,
        LedgerEntry.Category.DIVIDEND,
    ]

    cash_in = LedgerEntry.objects.filter(
        membership__sacco=sacco,
        category__in=cash_in_categories,
        entry_type=LedgerEntry.EntryType.CREDIT,
        created_at__date__lte=as_of_date,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    cash_out = LedgerEntry.objects.filter(
        membership__sacco=sacco,
        category__in=cash_out_categories,
        entry_type=LedgerEntry.EntryType.DEBIT,
        created_at__date__lte=as_of_date,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    # Note: ADJUSTMENT entries are excluded as they can be either direction
    # and require manual interpretation
    cash_balance = cash_in - cash_out

    # Liabilities: Member savings by type
    savings_by_type = {}

    for savings_type_name in SavingsType.Name.values:
        savings_total = Saving.objects.filter(
            membership__sacco=sacco,
            savings_type__name=savings_type_name,
            status=Saving.Status.ACTIVE,
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

        savings_by_type[savings_type_name] = str(savings_total.quantize(Decimal('0.01')))

    total_liabilities = sum(
        Decimal(amount) for amount in savings_by_type.values()
    )

    return {
        'as_of_date': as_of_date.isoformat(),
        'assets': {
            'loans_outstanding': str(loans_outstanding),
            'cash_balance': str(cash_balance),
            'total_assets': str(loans_outstanding + cash_balance),
        },
        'liabilities': {
            'savings_by_type': savings_by_type,
            'total_liabilities': str(total_liabilities),
        },
    }


def build_membership_return(sacco, period_start, period_end):
    """
    Build membership statistics return for SASRA reporting.

    Counts members by status, new registrations in period, and exits
    (LEFT status) in period.

    Args:
        sacco: Sacco instance
        period_start: date object for period start
        period_end: date object for period end

    Returns:
        dict with membership statistics
    """
    # Current membership by status
    current_memberships = Membership.objects.filter(sacco=sacco)

    members_by_status = {}
    for status in Membership.Status.values:
        count = current_memberships.filter(status=status).count()
        members_by_status[status] = count

    total_current = sum(members_by_status.values())

    # New registrations in period (approved during period)
    new_registrations = Membership.objects.filter(
        sacco=sacco,
        status=Membership.Status.APPROVED,
        approved_date__date__gte=period_start,
        approved_date__date__lte=period_end,
    ).count()

    # Exits in period (LEFT status with updated_at in period)
    exits = Membership.objects.filter(
        sacco=sacco,
        status=Membership.Status.LEFT,
        updated_at__date__gte=period_start,
        updated_at__date__lte=period_end,
    ).count()

    return {
        'period_start': period_start.isoformat(),
        'period_end': period_end.isoformat(),
        'current_members_by_status': members_by_status,
        'total_current_members': total_current,
        'new_registrations': new_registrations,
        'exits': exits,
    }


class SASRAReturnView(APIView):
    """
    API endpoint for generating SASRA regulatory returns.

    Supports PAR, financial position, and membership returns in JSON or XLSX format.
    """
    permission_classes = [IsAuthenticated, IsSaccoAdmin]

    def get(self, request):
        """Generate SASRA return based on query parameters."""
        report_type = request.query_params.get('type')
        as_of_date_str = request.query_params.get('as_of_date')
        period_start_str = request.query_params.get('period_start')
        period_end_str = request.query_params.get('period_end')
        output_format = request.query_params.get('format', 'json')

        # Validate report type
        valid_types = ['par', 'financial_position', 'membership']
        if report_type not in valid_types:
            return Response(
                {'error': f'Invalid report type. Must be one of: {valid_types}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parse dates
        if as_of_date_str:
            try:
                as_of_date = date.fromisoformat(as_of_date_str)
            except ValueError:
                return Response(
                    {'error': 'Invalid as_of_date format. Use YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            as_of_date = timezone.localdate()

        # For membership report, parse period dates
        if report_type == 'membership':
            if period_start_str:
                try:
                    period_start = date.fromisoformat(period_start_str)
                except ValueError:
                    return Response(
                        {'error': 'Invalid period_start format. Use YYYY-MM-DD.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            else:
                period_start = as_of_date.replace(day=1)

            if period_end_str:
                try:
                    period_end = date.fromisoformat(period_end_str)
                except ValueError:
                    return Response(
                        {'error': 'Invalid period_end format. Use YYYY-MM-DD.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            else:
                period_end = as_of_date

        # Get SACCO from header
        sacco_id = request.headers.get('X-Sacco-ID')
        if not sacco_id:
            return Response(
                {'error': 'X-Sacco-ID header is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from accounts.models import Sacco

        try:
            sacco = Sacco.objects.get(id=sacco_id)
        except Sacco.DoesNotExist:
            return Response(
                {'error': 'SACCO not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Verify user is admin for this SACCO
        if not request.user.roles.filter(
            name=Role.SACCO_ADMIN,
            sacco=sacco,
        ).exists():
            return Response(
                {'error': 'You are not an admin for this SACCO.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Generate report
        if report_type == 'par':
            report_data = build_par_return(sacco, as_of_date)
        elif report_type == 'financial_position':
            report_data = build_financial_position_return(sacco, as_of_date)
        else:  # membership
            report_data = build_membership_return(sacco, period_start, period_end)

        # Create audit log
        SystemAuditLog.objects.create(
            user=request.user,
            action='SASRA_RETURN_GENERATED',
            resource_type='SASRA_RETURN',
            resource_id=f'{report_type}_{as_of_date.isoformat()}',
            new_values={
                'report_type': report_type,
                'as_of_date': as_of_date.isoformat(),
                'sacco_id': str(sacco.id),
            },
        )

        # Return based on format
        if output_format == 'xlsx':
            return self._generate_xlsx(report_data, report_type, sacco, as_of_date)
        else:
            return Response(report_data)

    def _generate_xlsx(self, report_data, report_type, sacco, as_of_date):
        """Generate XLSX workbook from report data."""
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.upper()

        # Define headers based on report type
        if report_type == 'par':
            headers = [
                'Category',
                'Label',
                'Loan Count',
                'Outstanding Balance',
                'Provision Required',
            ]
            rows = []
            for cat_key, cat_data in report_data['categories'].items():
                rows.append([
                    cat_key,
                    cat_data['label'],
                    cat_data['loan_count'],
                    cat_data['outstanding_balance'],
                    cat_data['provision_required'],
                ])
            # Add totals row
            rows.append([
                'TOTAL',
                '',
                sum(r[2] for r in rows),
                report_data['total_outstanding_book'],
                str(sum(Decimal(r[4]) for r in rows)),
            ])
            # Add PAR ratios
            rows.append([])
            rows.append(['PAR30 Ratio', report_data['par30_ratio']])
            rows.append(['PAR90 Ratio', report_data['par90_ratio']])

        elif report_type == 'financial_position':
            headers = ['Item', 'Amount']
            rows = [
                ['ASSETS', ''],
                ['Loans Outstanding', report_data['assets']['loans_outstanding']],
                ['Cash Balance', report_data['assets']['cash_balance']],
                ['Total Assets', report_data['assets']['total_assets']],
                [],
                ['LIABILITIES', ''],
                ['BOSA Savings', report_data['liabilities']['savings_by_type'].get('BOSA', '0')],
                ['FOSA Savings', report_data['liabilities']['savings_by_type'].get('FOSA', '0')],
                ['Share Capital', report_data['liabilities']['savings_by_type'].get('SHARE_CAPITAL', '0')],
                ['Total Liabilities', report_data['liabilities']['total_liabilities']],
            ]

        else:  # membership
            headers = ['Status', 'Count']
            rows = []
            for status, count in report_data['current_members_by_status'].items():
                rows.append([status, count])
            rows.append(['Total Current', report_data['total_current_members']])
            rows.append([])
            rows.append(['New Registrations', report_data['new_registrations']])
            rows.append(['Exits', report_data['exits']])

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)

        # Write data rows
        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        filename = f'SASRA_{report_type}_{sacco.name}_{as_of_date.isoformat()}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

