"""Utilities for SACCO member CSV/Excel import jobs."""

import csv
import io
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from accounts.models import User
from saccomembership.models import Membership

from .models import MemberImportJob


MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
ALLOWED_EXTENSIONS = {'.csv', '.xlsx'}
REQUIRED_FIELDS = ('first_name', 'last_name', 'email')
OPTIONAL_FIELDS = ('phone_number', 'employment_status', 'monthly_income')
ALL_FIELDS = REQUIRED_FIELDS + OPTIONAL_FIELDS
HEADER_ALIASES = {
    'phone': 'phone_number',
    'phone number': 'phone_number',
    'employment status': 'employment_status',
    'monthly income': 'monthly_income',
}


def parse_import_file(uploaded_file):
    """
    Parse a CSV or XLSX member import file.

    Returns:
        tuple[list[dict], str | None]: Parsed rows and optional error message.
    """
    filename = (getattr(uploaded_file, 'name', '') or '').lower()
    extension = _file_extension(filename)

    if extension not in ALLOWED_EXTENSIONS:
        return [], 'Only .csv and .xlsx files are supported.'

    if getattr(uploaded_file, 'size', 0) > MAX_IMPORT_FILE_SIZE:
        return [], 'Import file must be smaller than 5MB.'

    try:
        if extension == '.csv':
            rows = _parse_csv(uploaded_file)
        else:
            rows = _parse_xlsx(uploaded_file)
    except Exception as exc:
        return [], f'Unable to read import file: {exc}'

    if not rows:
        return [], 'Import file contains no data rows.'

    return rows, None


def process_import_job(job_id, rows=None):
    """
    Process all rows for one MemberImportJob synchronously.

    TODO: Wrap this function in a Celery task for large imports.
    """
    job = MemberImportJob.objects.select_related('sacco', 'created_by').get(
        id=job_id,
    )
    job.status = MemberImportJob.Status.PROCESSING
    job.started_at = timezone.now()
    job.save(update_fields=['status', 'started_at'])

    if rows is None:
        return job

    job.total_rows = len(rows)
    job.processed_rows = 0
    job.success_rows = 0
    job.error_rows = 0
    job.errors = []
    job.save(
        update_fields=[
            'total_rows',
            'processed_rows',
            'success_rows',
            'error_rows',
            'errors',
        ],
    )

    for row_index, row in enumerate(rows, start=1):
        try:
            _import_member_row(row, job)
            job.success_rows += 1
        except Exception as exc:
            job.error_rows += 1
            job.errors.append(
                {
                    'row': row_index,
                    'field': _error_field_for_exception(exc, row),
                    'error': str(exc),
                },
            )
        job.processed_rows += 1

    if job.success_rows == 0 and job.error_rows > 0:
        job.status = MemberImportJob.Status.FAILED
    else:
        job.status = MemberImportJob.Status.COMPLETED

    job.completed_at = timezone.now()
    job.save(
        update_fields=[
            'status',
            'processed_rows',
            'success_rows',
            'error_rows',
            'errors',
            'completed_at',
        ],
    )
    return job


def _file_extension(filename):
    if filename.endswith('.xlsx'):
        return '.xlsx'
    if filename.endswith('.csv'):
        return '.csv'
    return ''


def _normalize_header(header):
    normalized = str(header or '').strip().lower()
    return HEADER_ALIASES.get(normalized, normalized.replace(' ', '_'))


def _normalize_row(raw_row):
    normalized = {}
    for key, value in raw_row.items():
        field_name = _normalize_header(key)
        if field_name not in ALL_FIELDS:
            continue
        if value is None:
            normalized[field_name] = None
            continue
        normalized[field_name] = str(value).strip()
    return normalized


def _parse_csv(uploaded_file):
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    content = uploaded_file.read()
    if isinstance(content, bytes):
        text_stream = io.StringIO(content.decode('utf-8-sig'))
    else:
        text_stream = io.StringIO(content)

    reader = csv.DictReader(text_stream)
    return [_normalize_row(row) for row in reader if any(row.values())]


def _parse_xlsx(uploaded_file):
    if hasattr(uploaded_file, 'seek'):
        uploaded_file.seek(0)

    workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        workbook.close()
        return []

    field_names = [_normalize_header(header) for header in headers]
    parsed_rows = []
    for values in rows_iter:
        if not any(values):
            continue
        raw_row = dict(zip(field_names, values))
        parsed_rows.append(_normalize_row(raw_row))

    workbook.close()
    return parsed_rows


def _validate_row(row):
    missing = [
        field for field in REQUIRED_FIELDS
        if not row.get(field)
    ]
    if missing:
        raise ValueError(
            f'Missing required fields: {", ".join(missing)}.',
        )

    monthly_income = row.get('monthly_income')
    if monthly_income:
        try:
            Decimal(str(monthly_income))
        except (InvalidOperation, ValueError) as exc:
            raise ValueError('monthly_income must be a valid number.') from exc


@transaction.atomic
def _import_member_row(row, job):
    _validate_row(row)

    user, created = User.objects.get_or_create(
        email=row['email'].lower(),
        defaults={
            'first_name': row['first_name'],
            'last_name': row['last_name'],
            'phone_number': row.get('phone_number') or None,
        },
    )
    if created:
        user.set_unusable_password()
        user.save(update_fields=['password'])

    membership_defaults = {'status': Membership.Status.APPROVED}
    if row.get('employment_status'):
        membership_defaults['notes'] = (
            f'Employment status: {row["employment_status"]}'
        )

    Membership.objects.update_or_create(
        user=user,
        sacco=job.sacco,
        defaults=membership_defaults,
    )


def _error_field_for_exception(exc, row):
    message = str(exc).lower()
    for field in REQUIRED_FIELDS + ('monthly_income',):
        if field in message:
            return field
    if 'email' in message:
        return 'email'
    return next(
        (field for field in REQUIRED_FIELDS if not row.get(field)),
        'row',
    )
